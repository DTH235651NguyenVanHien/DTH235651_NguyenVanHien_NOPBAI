import openpyxl
from openpyxl import Workbook


class NhanVien:
    def __init__(self, ma, ten, tuoi):
        self.ma = ma
        self.ten = ten
        self.tuoi = tuoi

    def __str__(self):
        return f"{self.ma} - {self.ten} - {self.tuoi}"


class QuanLyNhanVien:
    def __init__(self):
        self.nhanviens = []

    def them_nv(self, ma, ten, tuoi):
        self.nhanviens.append(NhanVien(ma, ten, tuoi))

    def sapxep_theo_tuoi(self):
        self.nhanviens.sort(key=lambda nv: nv.tuoi)

    def luu_excel(self, filename="nhanvien.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "DanhSachNV"

        # Ghi tiêu đề
        ws.append(["Mã", "Tên", "Tuổi"])

        # Ghi dữ liệu
        for nv in self.nhanviens:
            ws.append([nv.ma, nv.ten, nv.tuoi])

        wb.save(filename)
        print(f"Đã lưu danh sách nhân viên vào file {filename}")

    def doc_excel(self, filename="nhanvien.xlsx"):
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            self.nhanviens.clear()

            for row in ws.iter_rows(min_row=2, values_only=True):  # bỏ dòng tiêu đề
                ma, ten, tuoi = row
                self.nhanviens.append(NhanVien(ma, ten, tuoi))

            print(f"Đã đọc {len(self.nhanviens)} nhân viên từ file {filename}")
        except FileNotFoundError:
            print("Chưa có file Excel dữ liệu")


# ===================== DEMO =====================

if __name__ == "__main__":
    ql = QuanLyNhanVien()

    # Thêm nhân viên
    ql.them_nv("NV01", "Nguyễn Văn A", 25)
    ql.them_nv("NV02", "Trần Thị B", 30)
    ql.them_nv("NV03", "Lê Văn C", 22)

    # Lưu file Excel
    ql.luu_excel()

    # Đọc lại file Excel
    ql2 = QuanLyNhanVien()
    ql2.doc_excel()

    # In danh sách nhân viên
    print("\nDanh sách nhân viên:")
    for nv in ql2.nhanviens:
        print(nv)

    # Sắp xếp theo tuổi
    ql2.sapxep_theo_tuoi()
    print("\nDanh sách nhân viên theo tuổi tăng dần:")
    for nv in ql2.nhanviens:
        print(nv)


import openpyxl
from openpyxl import Workbook

# ====== DỮ LIỆU ======
nhanviens = []  # list để lưu tất cả nhân viên


# ====== HÀM QUẢN LÝ ======
def them_nv(ma, ten, tuoi):
    nv = {"ma": ma, "ten": ten, "tuoi": tuoi}
    nhanviens.append(nv)


def sapxep_theo_tuoi():
    nhanviens.sort(key=lambda x: x["tuoi"])


def luu_excel(filename="nhanvien.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "DanhSachNV"

    # Ghi tiêu đề
    ws.append(["Mã", "Tên", "Tuổi"])

    # Ghi dữ liệu
    for nv in nhanviens:
        ws.append([nv["ma"], nv["ten"], nv["tuoi"]])

    wb.save(filename)
    print(f"Đã lưu vào file {filename}")


def doc_excel(filename="nhanvien.xlsx"):
    global nhanviens
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        nhanviens = []

        # Bỏ dòng tiêu đề
        for row in ws.iter_rows(min_row=2, values_only=True):
            ma, ten, tuoi = row
            nhanviens.append({"ma": ma, "ten": ten, "tuoi": tuoi})

        print(f"Đã đọc {len(nhanviens)} nhân viên từ file {filename}")
    except FileNotFoundError:
        print("Chưa có file Excel")


def in_ds():
    for nv in nhanviens:
        print(nv)


# ====== DEMO ======
if __name__ == "__main__":
    # Thêm nhân viên
    them_nv("NV01", "Nguyễn Văn A", 25)
    them_nv("NV02", "Trần Thị B", 30)
    them_nv("NV03", "Lê Văn C", 22)

    # In danh sách
    print("\nDanh sách nhân viên ban đầu:")
    in_ds()

    # Lưu xuống file Excel
    luu_excel()

    # Đọc lại từ file
    doc_excel()
    print("\nDanh sách nhân viên sau khi đọc file:")
    in_ds()

    # Sắp xếp theo tuổi
    sapxep_theo_tuoi()
    print("\nDanh sách nhân viên theo tuổi tăng dần:")
    in_ds()
